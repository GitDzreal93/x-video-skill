#!/usr/bin/env python3
"""
AI漫剧分镜表导出脚本
将 JSON 格式的分镜数据导出为格式化的 Excel 文件
包含：角色三视图 Sheet + 分镜表（含首帧/尾帧）
"""

import json
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# 样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 场景交替颜色
SCENE_FILL_1 = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
SCENE_FILL_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# 分镜表列宽定义（含首帧/尾帧）
STORYBOARD_WIDTHS = {
    "A": 28,   # 拍摄场景
    "B": 6,    # 镜号
    "C": 30,   # 画面内容
    "D": 8,    # 景别
    "E": 18,   # 音乐/音效
    "F": 22,   # 台词
    "G": 6,    # 时长
    "H": 38,   # 首帧文生图提示词
    "I": 38,   # 尾帧文生图提示词
    "J": 38,   # 文生图提示词
    "K": 28,   # 图生视频提示词
}

STORYBOARD_HEADERS = [
    "拍摄场景", "镜号", "画面内容", "景别", "音乐/音效", "台词", "时长",
    "首帧文生图提示词", "尾帧文生图提示词", "文生图提示词", "图生视频提示词"
]

# 角色三视图列宽
CHARACTER_WIDTHS = {
    "A": 12,   # 角色名
    "B": 60,   # 三视图文生图提示词
    "C": 20,   # 标志性特征
    "D": 12,   # 标志色
}

CHARACTER_HEADERS = ["角色名", "三视图文生图提示词", "标志性特征", "标志色"]


def apply_cell_style(cell, font=CELL_FONT, alignment=CELL_ALIGN, border=THIN_BORDER, fill=None):
    """统一设置单元格样式"""
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill


def create_character_sheet(wb, characters):
    """创建角色三视图 Sheet"""
    if not characters:
        return

    ws = wb.create_sheet("角色三视图")

    # 标题
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "角色三视图提示词"
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 表头
    for col_idx, header in enumerate(CHARACTER_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_cell_style(cell, font=HEADER_FONT, alignment=HEADER_ALIGN, fill=HEADER_FILL)
    ws.row_dimensions[2].height = 28

    # 数据行
    for idx, char in enumerate(characters):
        row = 3 + idx
        values = [
            char.get("name", ""),
            char.get("turnaround_prompt", ""),
            char.get("signature_feature", ""),
            char.get("signature_color", ""),
        ]
        fill = SCENE_FILL_1 if idx % 2 == 0 else SCENE_FILL_2
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_cell_style(cell, fill=fill)
            if col_idx in (1, 4):
                cell.alignment = CENTER_ALIGN
        ws.row_dimensions[row].height = 60

    # 列宽
    for col_letter, width in CHARACTER_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A3"


def create_storyboard_sheet(wb, metadata, shots):
    """创建分镜表 Sheet"""
    ws = wb.active
    ws.title = "分镜表"

    # 标题行
    title = f"分镜表：{metadata.get('title', '未命名')}"
    last_col = chr(ord("A") + len(STORYBOARD_HEADERS) - 1)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 副标题
    subtitle_parts = []
    if metadata.get("style"):
        subtitle_parts.append(f"风格：{metadata['style']}")
    if metadata.get("duration"):
        subtitle_parts.append(f"总时长：{metadata['duration']}")
    if metadata.get("aspect_ratio"):
        subtitle_parts.append(f"画幅：{metadata['aspect_ratio']}")

    if subtitle_parts:
        ws.merge_cells(f"A2:{last_col}2")
        sub_cell = ws["A2"]
        sub_cell.value = "  |  ".join(subtitle_parts)
        sub_cell.font = Font(name="微软雅黑", size=10, color="666666")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 25

    # 表头行
    header_row = 3
    for col_idx, header in enumerate(STORYBOARD_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        apply_cell_style(cell, font=HEADER_FONT, alignment=HEADER_ALIGN, fill=HEADER_FILL)
    ws.row_dimensions[header_row].height = 30

    # 列宽
    for col_letter, width in STORYBOARD_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 数据行
    current_scene = None
    scene_group = 0
    for idx, shot in enumerate(shots):
        row = header_row + 1 + idx
        scene = shot.get("scene", "")

        if scene != current_scene:
            current_scene = scene
            scene_group += 1
        fill = SCENE_FILL_1 if scene_group % 2 == 1 else SCENE_FILL_2

        values = [
            scene,
            shot.get("shot_number", idx + 1),
            shot.get("description", ""),
            shot.get("shot_type", ""),
            shot.get("sound", ""),
            shot.get("dialogue", ""),
            shot.get("duration", ""),
            shot.get("first_frame_prompt", ""),
            shot.get("last_frame_prompt", ""),
            shot.get("image_prompt", ""),
            shot.get("video_prompt", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            apply_cell_style(cell, fill=fill)
            if col_idx in (2, 4, 7):
                cell.alignment = CENTER_ALIGN

        ws.row_dimensions[row].height = 50

    ws.freeze_panes = f"A{header_row + 1}"


def create_stats_sheet(wb, shots):
    """创建统计 Sheet"""
    ws = wb.create_sheet("统计")
    total_shots = len(shots)
    total_duration = sum(parse_duration(s.get("duration", "0")) for s in shots)

    stats = [
        ("总镜头数", total_shots),
        ("总时长（秒）", total_duration),
        ("平均每镜头（秒）", round(total_duration / max(total_shots, 1), 1)),
    ]

    ws["A1"] = "统计项目"
    ws["B1"] = "数值"
    apply_cell_style(ws["A1"], font=HEADER_FONT, alignment=HEADER_ALIGN, fill=HEADER_FILL)
    apply_cell_style(ws["B1"], font=HEADER_FONT, alignment=HEADER_ALIGN, fill=HEADER_FILL)

    for i, (label, value) in enumerate(stats, 2):
        apply_cell_style(ws.cell(row=i, column=1, value=label))
        apply_cell_style(ws.cell(row=i, column=2, value=value))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def create_workbook(metadata, shots, characters):
    """创建完整工作簿"""
    wb = openpyxl.Workbook()

    # Sheet 1: 分镜表
    create_storyboard_sheet(wb, metadata, shots)

    # Sheet 2: 角色三视图
    create_character_sheet(wb, characters)

    # Sheet 3: 统计
    create_stats_sheet(wb, shots)

    return wb


def parse_duration(duration_str):
    """解析时长字符串，如 '2s' → 2, '2.5s' → 2.5"""
    try:
        return float(str(duration_str).replace("s", "").strip())
    except (ValueError, AttributeError):
        return 0


def main():
    parser = argparse.ArgumentParser(description="AI漫剧分镜表导出工具")
    parser.add_argument("--data", required=True, help="JSON 格式的分镜数据字符串")
    parser.add_argument("--file", help="JSON 数据文件路径（替代 --data）")
    parser.add_argument("--output", "-o", help="输出文件路径（默认：./分镜表.xlsx）")
    args = parser.parse_args()

    # 读取数据
    if args.file:
        with open(args.file, "r", encoding="utf-8") as f:
            data = json.load(f)
    elif args.data:
        data = json.loads(args.data)
    else:
        print("请提供 --data 或 --file 参数")
        sys.exit(1)

    metadata = data.get("metadata", {})
    shots = data.get("shots", [])
    characters = data.get("characters", [])

    if not shots:
        print("错误：分镜数据为空")
        sys.exit(1)

    # 创建工作簿
    wb = create_workbook(metadata, shots, characters)

    # 输出路径
    output = args.output or f"./分镜表_{metadata.get('title', '未命名')}.xlsx"
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(output_path))
    print(f"导出成功：{output_path}")
    print(f"总镜头数：{len(shots)}")
    print(f"总时长：{sum(parse_duration(s.get('duration', '0')) for s in shots)}s")
    print(f"角色三视图：{len(characters)}个")


if __name__ == "__main__":
    main()
